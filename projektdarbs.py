import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook

def run_selenium(product):
    service = Service()
    option = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=option)

    url = "https://www.salidzini.lv"
    driver.get(url)
    time.sleep(2)
    accept_cookies = driver.find_element(By.XPATH, "//button[text()='Piekrītu']")
    accept_cookies.click()
    accept_cookies2 = driver.find_element(By.XPATH, "//button[text()='Nē, rādīt neatbilstošu reklāmu']")
    accept_cookies2.click()
    find = driver.find_element(By.ID, "q")
    find.send_keys(product)
    find.submit()
    time.sleep(2)
    price_caption = driver.find_element(By.ID, "price_caption")
    price_caption.click()
    price_from_field = driver.find_element(By.ID, "proce_from_field")
    price_from_field.send_keys(cena)
    price_submit = driver.find_element(By.XPATH, '//input[@type="submit" and @value="Atlasīt"]')
    price_submit.click()
    time.sleep(2)
    offers = driver.find_elements(By.XPATH, '//div[@itemprop="offers"]')

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Cena")
    ws.cell(row=1, column=2, value="Nosaukums")
    ws.cell(row=1, column=3, value="Saite")    

    data = []
    for offer in offers[:20]:
        title = offer.find_element(By.XPATH, './/h2[@itemprop="name"]').text
        price = offer.find_element(By.XPATH, './/span[@itemprop="price"]').text
        link = offer.find_element(By.XPATH, './/a[@itemprop="url"]').get_attribute('href')
        data.append([price, title, link])
        ws.append([price, title, link])

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width       
    wb.save(f'{product}.xlsx')
    
product = input("Ievadi produkta nosaukumu: ")
cena = input("Ievadi cena no:")
run_selenium(product)
